"""
Backerkit XLSX parser and validator.

Parses a Backerkit XLSX export from a Kickstarter campaign, validates structure
+ data, and returns either a list of preview-ready sale records or a list of
error strings.

Unlike Amazon/Ingram formats (one row = one aggregate sale), Backerkit rows
represent individual backer orders. This parser "rolls up" rows into aggregate
sales records grouped by (month/year, book, format).
"""

import re
from collections import defaultdict
from datetime import datetime

import openpyxl

from .base_parser import ParseResult, SalesImportParser
from ..models import Book

_SUCCESSFUL_STATUSES = {"collected", "imported"}


class BackerkitXLSXParser(SalesImportParser):
    """Parser for Backerkit XLSX export files from Kickstarter campaigns."""

    DISTRIBUTOR_NAME = None  # Kickstarter sales have no distributor

    _ITEM_PATTERN = re.compile(r"^item(\d+)$", re.IGNORECASE)

    def parse_and_validate(self, file, **kwargs) -> ParseResult:
        """
        Parse and validate a Backerkit XLSX file.

        Args:
            file: An UploadedFile (from request.FILES).

        Returns:
            ParseResult. preview is non-empty only when errors is empty.
            warnings contains deduplicated unknown item tags and skipped row
            numbers, and may accompany a valid preview.
        """
        filename = file.name or "unknown"
        timestamp = self._capture_timestamp()
        metadata = {"filename": filename, "timestamp": timestamp}
        comment = f"Kickstarter: File='{filename}' ({timestamp})"

        # ── Step 1: Open workbook ──────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception:
            return ParseResult(
                errors=["File is not a valid XLSX file."],
                metadata=metadata,
            )

        # ── Step 2: Exactly one sheet ──────────────────────────────────────
        if len(wb.sheetnames) != 1:
            return ParseResult(
                errors=[
                    f"Expected exactly 1 sheet, found {len(wb.sheetnames)}: "
                    f"{', '.join(wb.sheetnames)}."
                ],
                metadata=metadata,
            )

        ws = wb.active

        # ── Step 3: Parse header row and build column map ──────────────────
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {}
        for idx, val in enumerate(header_row):
            if val is not None:
                col_map[str(val).strip()] = idx

        errors = []

        # Required fixed columns
        for required in ("Pledge Status", "Order Placed"):
            if required not in col_map:
                errors.append(f"Missing required column: '{required}'.")

        # Discover itemN/qtyN pairs dynamically from headers
        item_qty_pairs = []  # list of (item_col_idx, qty_col_idx, N_str)
        for header, idx in col_map.items():
            m = self._ITEM_PATTERN.match(header)
            if m:
                n = m.group(1)
                qty_key = f"qty{n}"
                if qty_key in col_map:
                    item_qty_pairs.append((idx, col_map[qty_key], n))
                else:
                    errors.append(
                        f"Found column 'item{n}' but no corresponding 'qty{n}' column."
                    )

        if not item_qty_pairs and not errors:
            errors.append("No item/qty column pairs found in the file.")

        if errors:
            return ParseResult(errors=errors, metadata=metadata)

        # Sort pairs by N so they're processed in order
        item_qty_pairs.sort(key=lambda x: int(x[2]))

        pledge_col = col_map["Pledge Status"]
        date_col = col_map["Order Placed"]

        # ── Step 4: Process data rows ──────────────────────────────────────
        # aggregates key: (month, year, book_id, format)
        # aggregates value: {"qty": int, "book": Book}
        aggregates = defaultdict(lambda: {"qty": 0, "book": None})
        unknown_tags = set()
        skipped_rows = []
        row_errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip entirely blank rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            # ── Pledge Status ──────────────────────────────────────────────
            pledge_raw = (
                str(row[pledge_col]).strip() if row[pledge_col] is not None else ""
            )
            if pledge_raw.lower() not in _SUCCESSFUL_STATUSES:
                skipped_rows.append(row_idx)
                continue

            # ── Order Placed date ──────────────────────────────────────────
            date_raw = (
                str(row[date_col]).strip() if row[date_col] is not None else ""
            )
            try:
                dt = datetime.strptime(date_raw, "%m/%d/%y")
                month, year = dt.month, dt.year
            except (ValueError, TypeError):
                row_errors.append(
                    f"Row {row_idx}: 'Order Placed' value '{date_raw}' is not a valid "
                    f"date (expected MM/DD/YY)."
                )
                continue

            # ── Process each itemN/qtyN pair ──────────────────────────────
            for item_col_idx, qty_col_idx, n in item_qty_pairs:
                item_raw = row[item_col_idx]
                if item_raw is None or str(item_raw).strip() == "":
                    continue  # No item in this slot for this backer

                tag = str(item_raw).strip()

                # Look up tag against book kickstarter fields
                book = Book.objects.filter(kickstarter_item_tag_ebook=tag).first()
                if book:
                    fmt = "ebook"
                else:
                    book = Book.objects.filter(kickstarter_item_tag_print=tag).first()
                    if book:
                        fmt = "print"

                if not book:
                    unknown_tags.add(tag)
                    continue

                # Validate qty — only required for matched book items
                qty_raw = row[qty_col_idx]
                try:
                    qty_float = float(qty_raw) if qty_raw is not None else None
                    if qty_float is None:
                        raise ValueError("missing")
                    qty = int(qty_float)
                    if qty <= 0 or qty != qty_float:
                        raise ValueError("not a positive integer")
                except (ValueError, TypeError):
                    row_errors.append(
                        f"Row {row_idx}: qty{n} value '{qty_raw}' for item '{tag}' "
                        f"is not a valid positive integer."
                    )
                    continue

                key = (month, year, book.id, fmt)
                aggregates[key]["qty"] += qty
                aggregates[key]["book"] = book

        if row_errors:
            return ParseResult(errors=row_errors, metadata=metadata)

        # ── Step 5: At least one record must result ────────────────────────
        if not aggregates:
            return ParseResult(
                errors=[
                    "No valid sales records could be created from this file "
                    "(all rows were skipped or contained no matching Kickstarter item tags)."
                ],
                metadata=metadata,
            )

        # ── Step 6: Build preview rows and run serializer validation ───────
        preview_rows = []
        serializer_errors = []

        for (month, year, book_id, fmt), agg in aggregates.items():
            book = agg["book"]
            qty = agg["qty"]
            date_str = f"{year}-{month:02d}"

            # Compute revenue and royalty for preview display.
            # The serializer will recompute these on finalization, using the
            # same formula: revenue = qty * (cover_price - print_cost).
            revenue = self._money(qty * (book.cover_price - book.print_cost))
            royalty = self._money(book.hand_sold_author_royalty_rate * revenue)

            payload = {
                "book": book.id,
                "date": date_str,
                "quantity": qty,
                "sale_source": "kickstarter",
                "format": fmt,
                "currency": "USD",
                "author_paid": False,
                "comment": comment,
            }

            line_ref = f"Proposed record (book='{book.title}', {date_str}, {fmt})"
            validated_data, serial_errs = self._validate_with_serializer(payload, line_ref)
            if serial_errs:
                serializer_errors.extend(serial_errs)
                continue

            preview_rows.append({
                "book": book.id,
                "book_title": book.title,
                "book_label": f"{book.title} ({book.isbn_13})",
                "author_name": book.author.name,
                "royalty_rate": str(book.hand_sold_author_royalty_rate),
                "publication_date": book.publication_date.strftime("%Y-%m"),
                "date": date_str,
                "quantity": qty,
                "sale_source": "kickstarter",
                "distributor": None,
                "format": fmt,
                "currency": "USD",
                "publisher_revenue": str(revenue),
                "publisher_revenue_original": str(revenue),
                "author_royalty": str(royalty),
                "author_paid": False,
                "comment": comment,
            })

        if serializer_errors:
            return ParseResult(errors=serializer_errors, metadata=metadata)

        # ── Step 7: Build warnings (non-blocking) ──────────────────────────
        warnings = []
        if skipped_rows:
            warnings.append(
                "Rows with unsuccessful pledge status (skipped): "
                + ", ".join(str(r) for r in sorted(skipped_rows))
                + "."
            )
        for tag in sorted(unknown_tags):
            warnings.append(
                f"Unknown Kickstarter item tag (not matched to any book): {tag}"
            )

        return ParseResult(
            preview=preview_rows,
            warnings=warnings,
            metadata=metadata,
        )

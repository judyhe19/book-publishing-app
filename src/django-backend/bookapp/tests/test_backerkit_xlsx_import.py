import io
import re
import pytest
from decimal import Decimal
from openpyxl import Workbook
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient

from bookapp.models import Book, Author, Sale

pytestmark = pytest.mark.django_db

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_xlsx(rows, headers=None):
    """
    Build an in-memory XLSX with a single sheet.

    rows: list of dicts or lists. If dicts, headers must be provided.
    headers: list of column header strings for row 1.
    Returns bytes.
    """
    wb = Workbook()
    ws = wb.active
    if headers:
        ws.append(headers)
    for row in rows:
        if isinstance(row, dict):
            ws.append([row.get(h) for h in headers])
        else:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


DEFAULT_HEADERS = [
    "Pledge Status", "Order Placed", "item1", "qty1", "price1", "item2", "qty2", "price2"
]


def make_book(
    isbn_13,
    title="Test Book",
    author_name="Test Author",
    cover_price="20.00",
    print_cost="10.00",
    hand_sold_royalty_rate="0.20",
    distributor_royalty_rate="0.10",
    pub_date="2020-01-01",
    ks_tag_ebook=None,
    ks_tag_print=None,
    released=True,
):
    author, _ = Author.objects.get_or_create(
        name=author_name,
        defaults={"email": f"{author_name.lower().replace(' ', '_')}@test.com"},
    )
    return Book.objects.create(
        title=title,
        publication_date=pub_date,
        isbn_13=isbn_13,
        author=author,
        cover_price=Decimal(cover_price),
        print_cost=Decimal(print_cost),
        hand_sold_author_royalty_rate=Decimal(hand_sold_royalty_rate),
        distributor_author_royalty_rate=Decimal(distributor_royalty_rate),
        kickstarter_item_tag_ebook=ks_tag_ebook,
        kickstarter_item_tag_print=ks_tag_print,
        released=released,
    )


def upload(client, xlsx_bytes, filename="backerkit_export.xlsx"):
    f = SimpleUploadedFile(
        filename,
        xlsx_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return client.post(
        "/api/sales/import-backerkit-xlsx/",
        {"file": f},
        format="multipart",
    )


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def user():
    return User.objects.create_user(username="bkuser", password="pass12345")


@pytest.fixture
def authed_client(api_client, user):
    api_client.force_authenticate(user=user)
    return api_client


@pytest.fixture
def hobbit(db):
    return make_book(
        isbn_13="9780000000001",
        title="The Hobbit",
        cover_price="20.00",
        print_cost="8.00",
        hand_sold_royalty_rate="0.20",
        ks_tag_ebook="ebook-the-hobbit",
        ks_tag_print="paperback-the-hobbit",
    )


@pytest.fixture
def two_towers(db):
    return make_book(
        isbn_13="9780000000002",
        title="The Two Towers",
        cover_price="22.00",
        print_cost="9.00",
        hand_sold_royalty_rate="0.20",
        ks_tag_ebook="ebook-the-two-towers",
        ks_tag_print="paperback-the-two-towers",
    )


@pytest.fixture
def all_systems_red(db):
    return make_book(
        isbn_13="9780000000003",
        title="All Systems Red",
        cover_price="18.00",
        print_cost="7.00",
        hand_sold_royalty_rate="0.25",
        ks_tag_ebook="ebook-all-systems-red",
        ks_tag_print="paperback-all-systems-red",
    )


# ---------------------------------------------------------------------------
# Happy-path tests
# ---------------------------------------------------------------------------

class TestBackerkitHappyPath:
    def test_single_row_single_item(self, authed_client, hobbit):
        """Single backer row with one book item produces one sales record."""
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", 3, 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200, resp.data

        preview = resp.data["preview"]
        assert len(preview) == 1
        row = preview[0]
        assert row["book_title"] == "The Hobbit"
        assert row["format"] == "ebook"
        assert row["date"] == "2025-12"
        assert row["quantity"] == 3
        assert row["sale_source"] == "kickstarter"
        assert row["currency"] == "USD"
        assert row["distributor"] is None
        assert row["author_paid"] is False

    def test_revenue_and_royalty_computation(self, authed_client, hobbit):
        """Revenue = qty * (cover_price - print_cost); royalty = revenue * hand_sold_rate."""
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", 5, 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        row = resp.data["preview"][0]
        # cover=20, print=8 → margin=12; qty=5 → revenue=60.00
        assert row["publisher_revenue"] == "60.00"
        # royalty = 60.00 * 0.20 = 12.00
        assert row["author_royalty"] == "12.00"

    def test_rollup_same_book_format_month(self, authed_client, hobbit):
        """Two rows with the same book/format/month are summed into one record."""
        xlsx = make_xlsx(
            [
                ["imported", "12/08/25", "ebook-the-hobbit", 1, 0.0, None, None, None],
                ["imported", "12/08/25", "ebook-the-hobbit", 2, 0.0, None, None, None],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        preview = resp.data["preview"]
        assert len(preview) == 1
        assert preview[0]["quantity"] == 3

    def test_rollup_different_months_produce_separate_records(self, authed_client, hobbit):
        """Same book/format in different months → separate sales records."""
        xlsx = make_xlsx(
            [
                ["imported", "12/08/25", "ebook-the-hobbit", 1, 0.0, None, None, None],
                ["imported", "11/22/25", "ebook-the-hobbit", 5, 0.0, None, None, None],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        preview = resp.data["preview"]
        assert len(preview) == 2
        dates = {r["date"] for r in preview}
        assert dates == {"2025-12", "2025-11"}

    def test_rollup_different_formats_produce_separate_records(self, authed_client, hobbit):
        """Same book in ebook and print → two separate records."""
        xlsx = make_xlsx(
            [
                ["imported", "12/08/25", "ebook-the-hobbit", 2, 0.0, "paperback-the-hobbit", 3, 0.0],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        preview = resp.data["preview"]
        assert len(preview) == 2
        formats = {r["format"] for r in preview}
        assert formats == {"ebook", "print"}

    def test_multiple_items_per_row(self, authed_client, hobbit, two_towers, all_systems_red):
        """One backer row with multiple items contributes to multiple records."""
        xlsx = make_xlsx(
            [
                ["imported", "12/08/25", "ebook-the-hobbit", 1, 0.0, "ebook-the-two-towers", 2, 0.0],
                ["imported", "12/08/25", "ebook-the-two-towers", 4, 0.0, "paperback-all-systems-red", 8, 0.0],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        preview = resp.data["preview"]
        assert len(preview) == 3

        by_title = {r["book_title"]: r for r in preview}
        assert by_title["The Hobbit"]["quantity"] == 1
        assert by_title["The Two Towers"]["quantity"] == 6   # 2 + 4 rolled up
        assert by_title["All Systems Red"]["quantity"] == 8

    def test_collected_status_accepted(self, authed_client, hobbit):
        """'collected' pledge status is treated as successful."""
        xlsx = make_xlsx(
            [["collected", "12/08/25", "ebook-the-hobbit", 1, 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        assert len(resp.data["preview"]) == 1

    def test_failed_rows_skipped_with_warning(self, authed_client, hobbit):
        """Rows with unsuccessful pledge status are skipped and reported in warnings."""
        xlsx = make_xlsx(
            [
                ["imported", "12/08/25", "ebook-the-hobbit", 2, 0.0, None, None, None],
                ["failed", "11/22/25", "ebook-the-hobbit", 5, 0.0, None, None, None],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        assert len(resp.data["preview"]) == 1
        assert any("unsuccessful pledge status" in w for w in resp.data["warnings"])
        # Row 3 (data rows start at 2) should be mentioned
        assert any("3" in w for w in resp.data["warnings"])

    def test_unknown_tags_in_warnings_not_blocking(self, authed_client, hobbit):
        """Unknown item tags appear in warnings but don't block a valid import."""
        xlsx = make_xlsx(
            [
                ["imported", "12/08/25", "ebook-the-hobbit", 1, 0.0, "sticker-hp", 1, 0.0],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        assert len(resp.data["preview"]) == 1
        assert any("sticker-hp" in w for w in resp.data["warnings"])

    def test_unknown_tags_deduplicated(self, authed_client, hobbit):
        """The same unknown tag in multiple rows appears only once in warnings."""
        xlsx = make_xlsx(
            [
                ["imported", "12/08/25", "ebook-the-hobbit", 1, 0.0, "sticker-hp", 1, 0.0],
                ["imported", "11/22/25", "ebook-the-hobbit", 2, 0.0, "sticker-hp", 1, 0.0],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        tag_warnings = [w for w in resp.data["warnings"] if "sticker-hp" in w]
        assert len(tag_warnings) == 1

    def test_price_column_ignored(self, authed_client, hobbit):
        """The priceN columns are present but do not affect the output."""
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", 4, 999.99, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        assert resp.data["preview"][0]["quantity"] == 4

    def test_columns_in_nonstandard_order(self, authed_client, hobbit):
        """Column identification uses header names, not positions."""
        # Swap order: item1/qty1 come after Order Placed in reverse
        headers = ["item1", "qty1", "price1", "Pledge Status", "Order Placed", "item2", "qty2", "price2"]
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(["ebook-the-hobbit", 3, 0.0, "imported", "12/08/25", None, None, None])
        buf = io.BytesIO()
        wb.save(buf)
        xlsx = buf.getvalue()

        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        assert resp.data["preview"][0]["quantity"] == 3

    def test_blank_rows_skipped(self, authed_client, hobbit):
        """Completely blank rows in the sheet are silently ignored."""
        wb = Workbook()
        ws = wb.active
        ws.append(DEFAULT_HEADERS)
        ws.append(["imported", "12/08/25", "ebook-the-hobbit", 1, 0.0, None, None, None])
        ws.append([None, None, None, None, None, None, None, None])  # blank
        ws.append(["imported", "12/08/25", "ebook-the-hobbit", 2, 0.0, None, None, None])
        buf = io.BytesIO()
        wb.save(buf)

        resp = upload(authed_client, buf.getvalue())
        assert resp.status_code == 200
        # Quantities rolled up: 1 + 2 = 3
        assert resp.data["preview"][0]["quantity"] == 3

    def test_comment_format(self, authed_client, hobbit):
        """Comment follows 'Kickstarter: File=\\'...\\'  (YYYY-MM-DD HH:MM:SS)' format."""
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", 1, 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx, filename="backerkit_dec2025.xlsx")
        assert resp.status_code == 200
        comment = resp.data["preview"][0]["comment"]
        assert re.search(
            r"^Kickstarter: File='backerkit_dec2025\.xlsx' \(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\)$",
            comment,
        ), f"Comment did not match expected format: {comment!r}"

    def test_spec_example(self, authed_client, hobbit, two_towers, all_systems_red):
        """Reproduce the exact example from the spec."""
        # Spec table:
        # imported  12/08/25  ebook-the-hobbit    1  0.0  ebook-the-two-towers      2  0.0
        # imported  12/08/25  ebook-the-two-towers 4  0.0  paperback-all-systems-red 8  0.0
        # imported  11/22/25  ebook-the-hobbit    16  0.0  sticker-hp                0.0 (ignored qty)
        # failed    11/22/25  paperback-all-systems-red 32  0.0
        xlsx = make_xlsx(
            [
                ["imported", "12/08/25", "ebook-the-hobbit", 1, 0.0, "ebook-the-two-towers", 2, 0.0],
                ["imported", "12/08/25", "ebook-the-two-towers", 4, 0.0, "paperback-all-systems-red", 8, 0.0],
                ["imported", "11/22/25", "ebook-the-hobbit", 16, 0.0, "sticker-hp", None, 0.0],
                ["failed", "11/22/25", "paperback-all-systems-red", 32, 0.0, None, None, None],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200, resp.data

        preview = resp.data["preview"]
        # Expected 4 records: Hobbit/ebook/Dec, TwoTowers/ebook/Dec, AllSystemsRed/print/Dec, Hobbit/ebook/Nov
        assert len(preview) == 4

        by_key = {(r["book_title"], r["format"], r["date"]): r for r in preview}
        assert by_key[("The Hobbit", "ebook", "2025-12")]["quantity"] == 1
        assert by_key[("The Two Towers", "ebook", "2025-12")]["quantity"] == 6   # 2 + 4
        assert by_key[("All Systems Red", "print", "2025-12")]["quantity"] == 8
        assert by_key[("The Hobbit", "ebook", "2025-11")]["quantity"] == 16

        # sticker-hp in unknown tag warnings
        assert any("sticker-hp" in w for w in resp.data["warnings"])
        # Row 5 (failed) in skipped row warnings
        assert any("5" in w for w in resp.data["warnings"])

    def test_preview_only_no_sales_saved(self, authed_client, hobbit):
        """The import endpoint is preview-only; no Sale records are created."""
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", 3, 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        assert Sale.objects.count() == 0


# ---------------------------------------------------------------------------
# Structural / column validation errors
# ---------------------------------------------------------------------------

class TestBackerkitStructuralErrors:
    def test_not_xlsx(self, authed_client):
        """A non-XLSX file is rejected."""
        f = SimpleUploadedFile("bad.xlsx", b"not an xlsx file", content_type="application/octet-stream")
        resp = authed_client.post(
            "/api/sales/import-backerkit-xlsx/",
            {"file": f},
            format="multipart",
        )
        assert resp.status_code == 400
        assert any("valid XLSX" in e for e in resp.data["errors"])

    def test_multiple_sheets_rejected(self, authed_client):
        """A workbook with more than one sheet is rejected."""
        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.create_sheet("Sheet2")
        buf = io.BytesIO()
        wb.save(buf)

        resp = upload(authed_client, buf.getvalue())
        assert resp.status_code == 400
        assert any("1 sheet" in e for e in resp.data["errors"])

    def test_missing_pledge_status_column(self, authed_client):
        """Missing 'Pledge Status' column is a blocking error."""
        headers = ["Order Placed", "item1", "qty1", "price1"]
        xlsx = make_xlsx([["12/08/25", "tag", 1, 0.0]], headers=headers)
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("Pledge Status" in e for e in resp.data["errors"])

    def test_missing_order_placed_column(self, authed_client):
        """Missing 'Order Placed' column is a blocking error."""
        headers = ["Pledge Status", "item1", "qty1", "price1"]
        xlsx = make_xlsx([["imported", "tag", 1, 0.0]], headers=headers)
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("Order Placed" in e for e in resp.data["errors"])

    def test_no_item_qty_pairs(self, authed_client):
        """A sheet with no itemN/qtyN columns is rejected."""
        headers = ["Pledge Status", "Order Placed", "price1"]
        xlsx = make_xlsx([["imported", "12/08/25", 0.0]], headers=headers)
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("No item/qty" in e for e in resp.data["errors"])

    def test_item_without_matching_qty(self, authed_client):
        """An itemN column without a corresponding qtyN column is an error."""
        headers = ["Pledge Status", "Order Placed", "item1", "price1"]
        xlsx = make_xlsx([["imported", "12/08/25", "tag", 0.0]], headers=headers)
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("qty1" in e for e in resp.data["errors"])

    def test_no_file(self, authed_client):
        """Request with no file attached is rejected."""
        resp = authed_client.post(
            "/api/sales/import-backerkit-xlsx/",
            {},
            format="multipart",
        )
        assert resp.status_code == 400
        assert any("No file" in e for e in resp.data["errors"])

    def test_unauthenticated_rejected(self, api_client, hobbit):
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", 1, 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        f = SimpleUploadedFile("test.xlsx", xlsx)
        resp = api_client.post(
            "/api/sales/import-backerkit-xlsx/",
            {"file": f},
            format="multipart",
        )
        assert resp.status_code in (401, 403)


# ---------------------------------------------------------------------------
# Row-level data validation errors
# ---------------------------------------------------------------------------

class TestBackerkitRowErrors:
    def test_invalid_date_format(self, authed_client, hobbit):
        """An 'Order Placed' value that isn't MM/DD/YY is a blocking error."""
        xlsx = make_xlsx(
            [["imported", "2025-12-08", "ebook-the-hobbit", 1, 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("Order Placed" in e and "MM/DD/YY" in e for e in resp.data["errors"])

    def test_invalid_qty_non_integer(self, authed_client, hobbit):
        """A non-integer qty for a matched item is a blocking error."""
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", "abc", 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("positive integer" in e for e in resp.data["errors"])

    def test_invalid_qty_zero(self, authed_client, hobbit):
        """A qty of zero for a matched item is a blocking error."""
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", 0, 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("positive integer" in e for e in resp.data["errors"])

    def test_invalid_qty_negative(self, authed_client, hobbit):
        """A negative qty for a matched item is a blocking error."""
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", -5, 0.0, None, None, None]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("positive integer" in e for e in resp.data["errors"])

    def test_all_rows_failed_pledge_status(self, authed_client, hobbit):
        """If every row has an unsuccessful pledge status, no records can be created."""
        xlsx = make_xlsx(
            [
                ["failed", "12/08/25", "ebook-the-hobbit", 1, 0.0, None, None, None],
                ["cancelled", "11/22/25", "ebook-the-hobbit", 2, 0.0, None, None, None],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("No valid sales records" in e for e in resp.data["errors"])

    def test_all_rows_unknown_tags(self, authed_client):
        """If no row produces a matched item, no records can be created."""
        xlsx = make_xlsx(
            [["imported", "12/08/25", "sticker-hp", 3, 0.0, "mug-logo", 1, 0.0]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert any("No valid sales records" in e for e in resp.data["errors"])

    def test_multiple_row_errors_all_reported(self, authed_client, hobbit):
        """All row-level errors in the file are reported, not just the first."""
        xlsx = make_xlsx(
            [
                ["imported", "bad-date", "ebook-the-hobbit", 1, 0.0, None, None, None],
                ["imported", "also-bad", "ebook-the-hobbit", 2, 0.0, None, None, None],
            ],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 400
        assert len(resp.data["errors"]) >= 2

    def test_invalid_qty_for_unknown_tag_does_not_error(self, authed_client, hobbit):
        """qty for an unknown (non-book) item tag is not validated — it's ignored."""
        # sticker-hp has qty=0.0 (invalid as int) but since it's unknown, no error
        xlsx = make_xlsx(
            [["imported", "12/08/25", "ebook-the-hobbit", 2, 0.0, "sticker-hp", 0.0, 0.0]],
            headers=DEFAULT_HEADERS,
        )
        resp = upload(authed_client, xlsx)
        assert resp.status_code == 200
        assert resp.data["preview"][0]["quantity"] == 2

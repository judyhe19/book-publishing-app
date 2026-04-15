import pytest
import io
from decimal import Decimal
from django.contrib.auth.models import User
from rest_framework.test import APIClient
import openpyxl

from bookapp.models import Author, Book, Sale

pytestmark = pytest.mark.django_db

import uuid


# ---------------------------------------------------------------------------
# Shared fixtures (same pattern as test_royalty_report.py)
# ---------------------------------------------------------------------------

@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def user():
    return User.objects.create_user(username="fin_rpt_user", password="pass12345")


@pytest.fixture
def authed_client(api_client, user):
    api_client.force_authenticate(user=user)
    return api_client


def make_author(name="Report Author"):
    email = f"{name.lower().replace(' ', '_')}_{uuid.uuid4().hex[:6]}@test.com"
    return Author.objects.create(name=name, email=email)


def make_book(*, isbn_13, title, author, released=True, series_name=None, series_position=None,
              cover_price="20.00", print_cost="10.00"):
    return Book.objects.create(
        title=title,
        publication_date="2000-01-01",
        isbn_13=isbn_13,
        author=author,
        distributor_author_royalty_rate=Decimal("0.10"),
        hand_sold_author_royalty_rate=Decimal("0.20"),
        cover_price=Decimal(cover_price),
        print_cost=Decimal(print_cost),
        released=released,
        series_name=series_name,
        series_position=series_position,
    )


def make_sale(book, date, quantity=10, revenue="100.00", royalty=None,
              source="distributor", distributor="Ingram Spark", fmt="print",
              paid=False, kenp=None):
    if royalty is None:
        if source in ("handsold", "kickstarter"):
            royalty = Decimal(revenue) * book.hand_sold_author_royalty_rate
        else:
            royalty = Decimal(revenue) * book.distributor_author_royalty_rate
    return Sale.objects.create(
        book=book,
        author=book.author,
        date=date,
        quantity=quantity,
        publisher_revenue=Decimal(revenue),
        author_royalty=royalty,
        sale_source=source,
        distributor=distributor if source == "distributor" else None,
        format=fmt,
        author_paid=paid,
        kenp=kenp,
    )


def _parse_xlsx(response):
    """Parse an HttpResponse containing XLSX data and return the openpyxl workbook."""
    assert response.status_code == 200
    assert "spreadsheetml" in response["Content-Type"]
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    return wb


# ===========================================================================
# All Authors Royalty Report
# ===========================================================================

class TestAllAuthorsRoyaltyReport:

    def test_basic_structure(self, authed_client):
        a1 = make_author("Alice")
        a2 = make_author("Bob")
        b1 = make_book(isbn_13="9780000001001", title="Book A", author=a1)
        b2 = make_book(isbn_13="9780000001002", title="Book B", author=a2)
        make_sale(b1, "2025-01-15", revenue="100.00")
        make_sale(b2, "2025-01-15", revenue="200.00")

        resp = authed_client.get("/api/reports/all-authors-royalty/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 2,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        assert ws.title == "Author Royalties"

        # Header row
        headers = [c.value for c in ws[1]]
        assert headers[0] == "Author"
        assert "2025 Q1" in headers
        assert "2025 Q2" in headers
        assert headers[-1] == "Total"

    def test_quarterly_aggregation(self, authed_client):
        author = make_author("Charlie")
        book = make_book(isbn_13="9780000001003", title="Book C", author=author)
        make_sale(book, "2025-01-15", revenue="100.00")  # Q1: royalty = 10
        make_sale(book, "2025-04-15", revenue="200.00")  # Q2: royalty = 20

        resp = authed_client.get("/api/reports/all-authors-royalty/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 2,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        # Row 2 = Charlie, Row 3 = Totals
        row_data = [c.value for c in ws[2]]
        assert row_data[0] == "Charlie"
        assert row_data[1] == 10.0   # Q1 royalty
        assert row_data[2] == 20.0   # Q2 royalty
        assert row_data[3] == 30.0   # Total

    def test_alphabetical_sorting(self, authed_client):
        a_z = make_author("Zara")
        a_a = make_author("Aaron")
        make_book(isbn_13="9780000001004", title="Book Z", author=a_z)
        make_book(isbn_13="9780000001005", title="Book A", author=a_a)

        resp = authed_client.get("/api/reports/all-authors-royalty/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        # Last row is "Total"
        assert names[-1] == "Total"
        author_names = names[:-1]
        assert author_names == sorted(author_names)

    def test_total_row(self, authed_client):
        a1 = make_author("Dan")
        a2 = make_author("Eve")
        b1 = make_book(isbn_13="9780000001006", title="Book D", author=a1)
        b2 = make_book(isbn_13="9780000001007", title="Book E", author=a2)
        make_sale(b1, "2025-01-15", revenue="100.00")  # royalty = 10
        make_sale(b2, "2025-01-15", revenue="200.00")  # royalty = 20

        resp = authed_client.get("/api/reports/all-authors-royalty/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        total_row = [c.value for c in ws[ws.max_row]]
        assert total_row[0] == "Total"
        assert total_row[1] == 30.0   # Q1 total: 10 + 20
        assert total_row[2] == 30.0   # Grand total

    def test_excludes_projected_sales(self, authed_client):
        author = make_author("Frank")
        released_book = make_book(isbn_13="9780000001008", title="Released", author=author, released=True)
        unreleased_book = make_book(isbn_13="9780000001009", title="Unreleased", author=author, released=False)
        make_sale(released_book, "2025-01-15", revenue="100.00")
        make_sale(unreleased_book, "2025-01-15", revenue="500.00")

        resp = authed_client.get("/api/reports/all-authors-royalty/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        # Only released book royalty (10.00), not unreleased (50.00)
        total_row = [c.value for c in ws[ws.max_row]]
        assert total_row[1] == 10.0

    def test_requires_auth(self, api_client):
        resp = api_client.get("/api/reports/all-authors-royalty/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        assert resp.status_code in (401, 403)

    def test_invalid_params(self, authed_client):
        # Missing params
        resp = authed_client.get("/api/reports/all-authors-royalty/")
        assert resp.status_code == 400

        # Invalid quarter
        resp = authed_client.get("/api/reports/all-authors-royalty/", {
            "start_year": 2025, "start_quarter": 5,
            "end_year": 2025, "end_quarter": 1,
        })
        assert resp.status_code == 400

    def test_filename_format(self, authed_client):
        resp = authed_client.get("/api/reports/all-authors-royalty/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        disposition = resp["Content-Disposition"]
        assert "All_Authors_Royalty_Report_" in disposition
        assert ".xlsx" in disposition


# ===========================================================================
# Publisher Profit Report
# ===========================================================================

class TestPublisherProfitReport:

    def test_basic_structure(self, authed_client):
        author = make_author("Grace")
        book = make_book(isbn_13="9780000002001", title="Profit Book", author=author)
        make_sale(book, "2025-01-15", revenue="100.00")

        resp = authed_client.get("/api/reports/publisher-profit/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        assert ws.title == "Publisher Profit Report"

        headers = [c.value for c in ws[1]]
        assert headers[0] == "Author"
        assert headers[1] == "Series/Position"
        assert headers[2] == "Title"
        assert headers[3] == "ISBN-13"
        assert headers[4] == "ASIN"
        assert headers[5] == "Cover Price"
        assert headers[6] == "Print Cost"
        assert "2025 Q1" in headers
        assert headers[-1] == "Total"

    def test_profit_calculation(self, authed_client):
        author = make_author("Henry")
        book = make_book(isbn_13="9780000002002", title="Profit Calc", author=author)
        # revenue=100, royalty=10, profit=90
        make_sale(book, "2025-01-15", revenue="100.00")

        resp = authed_client.get("/api/reports/publisher-profit/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        # Row 2 = book data
        row = [c.value for c in ws[2]]
        q1_col_idx = 7  # After Author, Series/Position, Title, ISBN-13, ASIN, Cover Price, Print Cost
        assert row[q1_col_idx] == 90.0  # profit = 100 - 10
        assert row[q1_col_idx + 1] == 90.0  # Total

    def test_excludes_unreleased_books(self, authed_client):
        author = make_author("Iris")
        released = make_book(isbn_13="9780000002003", title="Released Book", author=author, released=True)
        unreleased = make_book(isbn_13="9780000002004", title="Unreleased Book", author=author, released=False)
        make_sale(released, "2025-01-15", revenue="100.00")
        make_sale(unreleased, "2025-01-15", revenue="500.00")

        resp = authed_client.get("/api/reports/publisher-profit/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        # Only 2 data rows: header + released book + total
        assert ws.max_row == 3
        assert ws.cell(row=2, column=3).value == "Released Book"

    def test_sorting_order(self, authed_client):
        a_b = make_author("Bob Author")
        a_a = make_author("Alice Author")
        b2 = make_book(isbn_13="9780000002005", title="Zeta", author=a_b)
        b1 = make_book(isbn_13="9780000002006", title="Alpha", author=a_a)

        resp = authed_client.get("/api/reports/publisher-profit/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        # Alice Author should come before Bob Author
        assert ws.cell(row=2, column=1).value == "Alice Author"
        assert ws.cell(row=3, column=1).value == "Bob Author"

    def test_total_row(self, authed_client):
        author = make_author("Jake")
        b1 = make_book(isbn_13="9780000002007", title="Book J1", author=author)
        b2 = make_book(isbn_13="9780000002008", title="Book J2", author=author)
        make_sale(b1, "2025-01-15", revenue="100.00")  # profit = 90
        make_sale(b2, "2025-01-15", revenue="200.00")  # profit = 180

        resp = authed_client.get("/api/reports/publisher-profit/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        total_row = [c.value for c in ws[ws.max_row]]
        assert total_row[0] == "Total"
        # Series/Position, Title, ISBN, ASIN, Cover Price, Print Cost should be blank
        for i in range(1, 7):
            assert total_row[i] == "" or total_row[i] is None
        # Q1 profit total
        assert total_row[7] == 270.0  # 90 + 180

    def test_series_display(self, authed_client):
        author = make_author("Kate")
        book = make_book(isbn_13="9780000002009", title="Series Book", author=author,
                         series_name="Epic Saga", series_position=3)

        resp = authed_client.get("/api/reports/publisher-profit/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        wb = _parse_xlsx(resp)
        ws = wb.active

        assert ws.cell(row=2, column=2).value == "Epic Saga (3)"

    def test_filename_format(self, authed_client):
        resp = authed_client.get("/api/reports/publisher-profit/", {
            "start_year": 2025, "start_quarter": 1,
            "end_year": 2025, "end_quarter": 1,
        })
        disposition = resp["Content-Disposition"]
        assert "Publisher_Profit_Report_" in disposition
        assert ".xlsx" in disposition


# ===========================================================================
# Amazon Sales Report
# ===========================================================================

class TestAmazonSalesReport:

    def test_basic_structure(self, authed_client):
        author = make_author("Liam")
        book = make_book(isbn_13="9780000003001", title="Amazon Book", author=author,
                         cover_price="15.00", print_cost="5.00")
        make_sale(book, "2025-01-15", revenue="50.00", distributor="Amazon", fmt="print")

        resp = authed_client.get("/api/reports/amazon-sales/")
        wb = _parse_xlsx(resp)
        ws = wb.active

        assert ws.title == "Amazon Sales"

        headers = [c.value for c in ws[1]]
        expected = [
            "Author", "Series/Position", "Title", "ISBN-13", "ASIN",
            "Print Quantity", "Print Revenue",
            "Ebook Quantity", "Ebook Revenue",
            "KENP", "KENP Revenue",
        ]
        assert headers == expected

    def test_aggregates_amazon_only(self, authed_client):
        author = make_author("Mia")
        book = make_book(isbn_13="9780000003002", title="Multi Dist", author=author)
        # Amazon sale
        make_sale(book, "2025-01-15", quantity=10, revenue="100.00",
                  distributor="Amazon", fmt="print")
        # Ingram sale (should be excluded)
        make_sale(book, "2025-01-15", quantity=20, revenue="200.00",
                  distributor="Ingram Spark", fmt="print")

        resp = authed_client.get("/api/reports/amazon-sales/")
        wb = _parse_xlsx(resp)
        ws = wb.active

        # Print Quantity should only include Amazon
        assert ws.cell(row=2, column=6).value == 10

    def test_print_ebook_kenp_split(self, authed_client):
        author = make_author("Noah")
        book = make_book(isbn_13="9780000003003", title="Split Book", author=author)

        # Print sale
        make_sale(book, "2025-01-15", quantity=5, revenue="50.00",
                  distributor="Amazon", fmt="print")
        # Ebook sale
        make_sale(book, "2025-02-15", quantity=3, revenue="30.00",
                  distributor="Amazon", fmt="ebook")
        # KENP sale
        make_sale(book, "2025-03-15", quantity=None, revenue="15.00",
                  distributor="Amazon", fmt="kindle unlimited", kenp=1000)

        resp = authed_client.get("/api/reports/amazon-sales/")
        wb = _parse_xlsx(resp)
        ws = wb.active

        row = [c.value for c in ws[2]]
        # Print Qty, Print Rev, Ebook Qty, Ebook Rev, KENP, KENP Rev
        assert row[5] == 5      # Print Quantity
        assert row[6] == 50.0   # Print Revenue
        assert row[7] == 3      # Ebook Quantity
        assert row[8] == 30.0   # Ebook Revenue
        assert row[9] == 1000   # KENP
        assert row[10] == 15.0  # KENP Revenue

    def test_total_row(self, authed_client):
        author = make_author("Olivia")
        b1 = make_book(isbn_13="9780000003004", title="Book O1", author=author)
        b2 = make_book(isbn_13="9780000003005", title="Book O2", author=author)
        make_sale(b1, "2025-01-15", quantity=10, revenue="100.00",
                  distributor="Amazon", fmt="print")
        make_sale(b2, "2025-01-15", quantity=20, revenue="200.00",
                  distributor="Amazon", fmt="print")

        resp = authed_client.get("/api/reports/amazon-sales/")
        wb = _parse_xlsx(resp)
        ws = wb.active

        total_row = [c.value for c in ws[ws.max_row]]
        assert total_row[0] == "Total"
        # Series/Position, Title, ISBN, ASIN should be blank
        for i in range(1, 5):
            assert total_row[i] == "" or total_row[i] is None
        assert total_row[5] == 30    # Print Qty: 10 + 20
        assert total_row[6] == 300.0 # Print Rev: 100 + 200

    def test_excludes_projected(self, authed_client):
        author = make_author("Pam")
        released = make_book(isbn_13="9780000003006", title="Released", author=author, released=True)
        unreleased = make_book(isbn_13="9780000003007", title="Unreleased", author=author, released=False)
        make_sale(released, "2025-01-15", quantity=10, revenue="100.00",
                  distributor="Amazon", fmt="print")
        make_sale(unreleased, "2025-01-15", quantity=50, revenue="500.00",
                  distributor="Amazon", fmt="print")

        resp = authed_client.get("/api/reports/amazon-sales/")
        wb = _parse_xlsx(resp)
        ws = wb.active

        # Only released book + total row = 3 rows
        assert ws.max_row == 3
        assert ws.cell(row=2, column=3).value == "Released"

    def test_requires_auth(self, api_client):
        resp = api_client.get("/api/reports/amazon-sales/")
        assert resp.status_code in (401, 403)

    def test_no_params_needed(self, authed_client):
        """Amazon sales report does not require quarter params."""
        resp = authed_client.get("/api/reports/amazon-sales/")
        assert resp.status_code == 200

    def test_filename_format(self, authed_client):
        resp = authed_client.get("/api/reports/amazon-sales/")
        disposition = resp["Content-Disposition"]
        assert "Amazon_Sale_Report_" in disposition
        assert ".xlsx" in disposition

    def test_sorting_order(self, authed_client):
        a_b = make_author("Zack")
        a_a = make_author("Anna")
        make_book(isbn_13="9780000003008", title="Zeta", author=a_b)
        make_book(isbn_13="9780000003009", title="Alpha", author=a_a)

        resp = authed_client.get("/api/reports/amazon-sales/")
        wb = _parse_xlsx(resp)
        ws = wb.active

        # Anna should come before Zack
        assert ws.cell(row=2, column=1).value == "Anna"
        assert ws.cell(row=3, column=1).value == "Zack"

# views/financial_reports.py
# XLSX export views: All Authors Royalty, Publisher Profit, Amazon Sales

import io
from datetime import datetime
from decimal import Decimal

from django.db.models import (
    Sum, Case, When, Value, DecimalField, IntegerField, F,
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse

from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

import openpyxl

from ..models import Author, Book, Sale
from ..utils.quarters import quarter_date_range, enumerate_quarters, validate_quarter_params


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _get_released_books_queryset():
    """
    Return a queryset of released books sorted by:
    author name → series name → series position → title.
    """
    return (
        Book.objects
        .filter(released=True)
        .select_related("author")
        .order_by("author__name", "series_name", "series_position", "title")
    )


def _xlsx_response(workbook, filename):
    """Write an openpyxl workbook to an HttpResponse for file download."""
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _make_timestamp():
    """Return the current timestamp formatted for filenames: YYYY-MM-DD H_MM_SS."""
    now = datetime.now()
    return now.strftime("%Y-%m-%d %H_%M_%S")


# ---------------------------------------------------------------------------
# 1. All Authors Royalty Report
# ---------------------------------------------------------------------------

class AllAuthorsRoyaltyReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Validate quarter params
        result = validate_quarter_params(request)
        if isinstance(result, Response):
            return result
        start_year, start_quarter, end_year, end_quarter = result

        # Build quarter labels
        quarter_labels = []
        for y, q in enumerate_quarters(start_year, start_quarter, end_year, end_quarter):
            quarter_labels.append((y, q, f"{y} Q{q}"))

        # All authors, sorted alphabetically by name
        authors = Author.objects.all().order_by("name")

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Author Royalties"

        # Header row
        headers = ["Author"] + [label for _, _, label in quarter_labels] + ["Total"]
        ws.append(headers)

        # Per-column totals for the final Total row
        column_totals = [Decimal("0.00")] * len(quarter_labels)
        grand_total = Decimal("0.00")

        for author in authors:
            row = [author.name]
            author_total = Decimal("0.00")

            for idx, (y, q, _label) in enumerate(quarter_labels):
                sd, ed = quarter_date_range(y, q)
                royalty = (
                    Sale.objects
                    .filter(
                        book__author=author,
                        book__released=True,
                        date__gte=sd,
                        date__lte=ed,
                    )
                    .aggregate(
                        total=Coalesce(
                            Sum("author_royalty"),
                            Value(Decimal("0.00")),
                            output_field=DecimalField(max_digits=12, decimal_places=2),
                        )
                    )["total"]
                )
                row.append(float(royalty))
                author_total += royalty
                column_totals[idx] += royalty

            row.append(float(author_total))
            grand_total += author_total
            ws.append(row)

        # Total row
        total_row = ["Total"] + [float(t) for t in column_totals] + [float(grand_total)]
        ws.append(total_row)

        filename = f"All_Authors_Royalty_Report_{_make_timestamp()}.xlsx"
        return _xlsx_response(wb, filename)


# ---------------------------------------------------------------------------
# 2. Publisher Profit Report
# ---------------------------------------------------------------------------

class PublisherProfitReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Validate quarter params
        result = validate_quarter_params(request)
        if isinstance(result, Response):
            return result
        start_year, start_quarter, end_year, end_quarter = result

        # Build quarter labels
        quarter_labels = []
        for y, q in enumerate_quarters(start_year, start_quarter, end_year, end_quarter):
            quarter_labels.append((y, q, f"{y} Q{q}"))

        books = _get_released_books_queryset()

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Publisher Profit Report"

        # Header row
        headers = [
            "Author", "Series/Position", "Title", "ISBN-13", "ASIN",
            "Cover Price", "Print Cost",
        ] + [label for _, _, label in quarter_labels] + ["Total"]
        ws.append(headers)

        # Per-column totals for the final Total row
        num_static_cols = 7  # Author through Print Cost
        column_totals = [Decimal("0.00")] * len(quarter_labels)
        grand_total = Decimal("0.00")

        for book in books:
            series_display = f"{book.series_name} ({book.series_position})" if book.series_name else ""
            row = [
                book.author.name,
                series_display,
                book.title,
                book.isbn_13,
                book.amazon_asin_ebook or "",
                float(book.cover_price),
                float(book.print_cost),
            ]

            book_total = Decimal("0.00")
            for idx, (y, q, _label) in enumerate(quarter_labels):
                sd, ed = quarter_date_range(y, q)
                agg = (
                    Sale.objects
                    .filter(
                        book=book,
                        date__gte=sd,
                        date__lte=ed,
                    )
                    .aggregate(
                        revenue=Coalesce(
                            Sum("publisher_revenue"),
                            Value(Decimal("0.00")),
                            output_field=DecimalField(max_digits=12, decimal_places=2),
                        ),
                        royalty=Coalesce(
                            Sum("author_royalty"),
                            Value(Decimal("0.00")),
                            output_field=DecimalField(max_digits=12, decimal_places=2),
                        ),
                    )
                )
                profit = agg["revenue"] - agg["royalty"]
                row.append(float(profit))
                book_total += profit
                column_totals[idx] += profit

            row.append(float(book_total))
            grand_total += book_total
            ws.append(row)

        # Total row — static columns blank except for Author="Total"
        total_row = ["Total"] + [""] * (num_static_cols - 1)
        total_row += [float(t) for t in column_totals]
        total_row.append(float(grand_total))
        ws.append(total_row)

        filename = f"Publisher_Profit_Report_{_make_timestamp()}.xlsx"
        return _xlsx_response(wb, filename)


# ---------------------------------------------------------------------------
# 3. Amazon Sales Report
# ---------------------------------------------------------------------------

class AmazonSalesReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        books = _get_released_books_queryset()

        # Build workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Amazon Sales"

        # Header row
        headers = [
            "Author", "Series/Position", "Title", "ISBN-13", "ASIN",
            "Print Quantity", "Print Revenue",
            "Ebook Quantity", "Ebook Revenue",
            "KENP", "KENP Revenue",
        ]
        ws.append(headers)

        # Running totals for the final row
        totals = {
            "print_qty": 0,
            "print_rev": Decimal("0.00"),
            "ebook_qty": 0,
            "ebook_rev": Decimal("0.00"),
            "kenp": 0,
            "kenp_rev": Decimal("0.00"),
        }

        for book in books:
            series_display = f"{book.series_name} ({book.series_position})" if book.series_name else ""

            # Base queryset: Amazon distributor sales, non-projected
            amazon_sales = Sale.objects.filter(
                book=book,
                sale_source="distributor",
                distributor="Amazon",
            )

            # Print: format="print"
            print_agg = amazon_sales.filter(format="print").aggregate(
                qty=Coalesce(Sum("quantity"), Value(0), output_field=IntegerField()),
                rev=Coalesce(
                    Sum("publisher_revenue"),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
            )

            # Ebook: format="ebook"
            ebook_agg = amazon_sales.filter(format="ebook").aggregate(
                qty=Coalesce(Sum("quantity"), Value(0), output_field=IntegerField()),
                rev=Coalesce(
                    Sum("publisher_revenue"),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
            )

            # KENP: format="kindle unlimited"
            kenp_agg = amazon_sales.filter(format="kindle unlimited").aggregate(
                kenp=Coalesce(Sum("kenp"), Value(0), output_field=IntegerField()),
                rev=Coalesce(
                    Sum("publisher_revenue"),
                    Value(Decimal("0.00")),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                ),
            )

            row = [
                book.author.name,
                series_display,
                book.title,
                book.isbn_13,
                book.amazon_asin_ebook or "",
                print_agg["qty"],
                float(print_agg["rev"]),
                ebook_agg["qty"],
                float(ebook_agg["rev"]),
                kenp_agg["kenp"],
                float(kenp_agg["rev"]),
            ]
            ws.append(row)

            # Accumulate totals
            totals["print_qty"] += print_agg["qty"]
            totals["print_rev"] += print_agg["rev"]
            totals["ebook_qty"] += ebook_agg["qty"]
            totals["ebook_rev"] += ebook_agg["rev"]
            totals["kenp"] += kenp_agg["kenp"]
            totals["kenp_rev"] += kenp_agg["rev"]

        # Total row
        total_row = [
            "Total", "", "", "", "",
            totals["print_qty"],
            float(totals["print_rev"]),
            totals["ebook_qty"],
            float(totals["ebook_rev"]),
            totals["kenp"],
            float(totals["kenp_rev"]),
        ]
        ws.append(total_row)

        filename = f"Amazon_Sale_Report_{_make_timestamp()}.xlsx"
        return _xlsx_response(wb, filename)

"""
Amazon XLSX parser and validator.

Parses an Amazon monthly sales XLSX, validates structure + data across all
supported sheets, and returns preview-ready sale records or error/warning lists.
"""

import datetime
from abc import ABC, abstractmethod
from decimal import Decimal, InvalidOperation

import openpyxl

from .base_parser import ParseResult, SalesImportParser
from .currency_converter import CurrencyConverter, CurrencyConversionError


# ──────────────────────────────────────────────────────────────────────────────
# Module-level helpers (shared by sheet handlers)
# ──────────────────────────────────────────────────────────────────────────────

def _parse_sales_period(cell_value) -> tuple[int, int]:
    """Parse 'June 2025' → (6, 2025). Raises ValueError if format is not recognised."""
    dt = datetime.datetime.strptime(str(cell_value).strip(), "%B %Y")
    return dt.month, dt.year


def _build_col_map(header_row) -> dict[str, int]:
    """Return {column_name: 0-based_index} for a row of cell values, skipping blank cells."""
    return {
        str(v).strip(): i
        for i, v in enumerate(header_row)
        if v is not None and str(v).strip()
    }


def _cell(row: tuple, col_map: dict[str, int], col_name: str) -> str:
    """Return the stripped string value of a cell by column name, or '' if absent/None."""
    idx = col_map.get(col_name)
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _is_blank_row(row: tuple) -> bool:
    """True if every cell in the row is None or an empty/whitespace string."""
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row)


# ──────────────────────────────────────────────────────────────────────────────
# Sheet-level abstraction (private to this module)
# ──────────────────────────────────────────────────────────────────────────────

class _SheetHandler(ABC):
    """
    Internal abstraction for a single Amazon XLSX sheet type.

    Each supported sheet (Paperback Royalty, Hardcover Royalty, eBook Royalty,
    KENP) has its own column schema, book-lookup key, and field mapping.
    Subclasses encapsulate that per-sheet logic so AmazonXLSXParser.parse_and_validate
    can iterate over handlers without branching on sheet name.

    Not a subclass of SalesImportParser — it is an internal detail of
    AmazonXLSXParser. The parent parser instance is passed to parse_rows so
    handlers can call shared helpers (_lookup_book_by_isbn, _money, etc.) and
    access parser._converter for currency conversion.
    """

    #: The exact sheet name as it appears in the XLSX file.
    sheet_name: str

    @abstractmethod
    def parse_rows(
        self,
        worksheet,
        filename: str,
        timestamp: str,
        parser: "AmazonXLSXParser",
    ) -> tuple[list[dict], list[str], list[str]]:
        """
        Parse all data rows from a single worksheet.

        Args:
            worksheet: An openpyxl Worksheet object.
            filename:  Original upload filename (for comment generation).
            timestamp: Import timestamp string (for comment generation).
            parser:    The owning AmazonXLSXParser instance. Use its protected
                       helpers (_lookup_book_by_isbn, _lookup_book_by_asin,
                       _compute_author_royalty, _money) and parser._converter
                       for currency conversion.

        Returns:
            (preview_rows, errors, warnings)
            - preview_rows: list of preview row dicts (empty if any errors).
            - errors:       blocking validation errors for this sheet.
            - warnings:     non-blocking notices (e.g. skipped unsupported rows).
        """
        ...

    # ──────────────────────────────────────────────────────────────────────────
    # Protected helpers shared by concrete handlers
    # ──────────────────────────────────────────────────────────────────────────

    def _extract_header(self, worksheet) -> tuple:
        """
        Read the first two rows of the worksheet:
          - Row 1: validate A1 == 'Sales Period' and parse B1 as 'Month YYYY'.
          - Row 2: build a column-name → 0-based-index map.

        Returns ((month, year), col_map, errors). On any error, (month, year)
        and col_map are None and errors is non-empty.
        """
        a1 = worksheet.cell(row=1, column=1).value
        b1 = worksheet.cell(row=1, column=2).value

        if (str(a1).strip() if a1 is not None else "") != "Sales Period":
            return None, None, [
                f"Sheet '{self.sheet_name}': Expected 'Sales Period' in A1, got '{a1}'."
            ]

        try:
            month, year = _parse_sales_period(b1)
        except (ValueError, TypeError):
            return None, None, [
                f"Sheet '{self.sheet_name}': Cannot parse sales period from B1 "
                f"(got '{b1}'). Expected format: 'June 2025'."
            ]

        header_row = tuple(cell.value for cell in worksheet[2])
        col_map = _build_col_map(header_row)
        return (month, year), col_map, []

    def _check_required_cols(self, col_map: dict, required: set) -> list[str]:
        """Return error strings for any required columns absent from col_map."""
        missing = required - col_map.keys()
        if missing:
            return [
                f"Sheet '{self.sheet_name}': Missing required columns: "
                f"{', '.join(sorted(missing))}."
            ]
        return []

    def _convert_royalty(
        self, royalty_raw: str, currency_raw: str, row_ref: str, parser
    ) -> tuple:
        """
        Parse royalty_raw as a Decimal and convert to USD via parser._converter.

        Returns (royalty_original, royalty_usd, errors). On any failure,
        royalty_original and royalty_usd are None and errors is non-empty.
        """
        if not royalty_raw:
            return None, None, [f"{row_ref}: Royalty cannot be empty."]
        if not currency_raw:
            return None, None, [f"{row_ref}: Currency cannot be empty."]

        try:
            royalty_original = Decimal(royalty_raw)
        except InvalidOperation:
            return None, None, [
                f"{row_ref}: Royalty '{royalty_raw}' is not a valid number."
            ]

        try:
            royalty_usd = parser._converter.to_usd(royalty_original, currency_raw)
        except CurrencyConversionError as exc:
            return None, None, [f"{row_ref}: Currency conversion failed — {exc}"]

        return royalty_original, royalty_usd, []

    def _make_comment(
        self, marketplace: str, currency: str, filename: str, timestamp: str
    ) -> str:
        mkt = marketplace[:47] + "..." if len(marketplace) > 50 else marketplace
        fname = filename[:153] + "..." if len(filename) > 156 else filename
        return (
            f"Amazon: Sheet='{self.sheet_name}' Marketplace='{mkt}' "
            f"Currency='{currency}' File='{fname}' ({timestamp})"
        )

    def _preview_row_base(self, book, sale_date_str: str, parser) -> dict:
        """Return the fields shared by all Amazon preview row dicts."""
        return {
            "book": book.id,
            "book_title": book.title,
            "book_label": f"{book.title} ({book.isbn_13})",
            "author_name": book.author.name,
            "royalty_rate": str(book.distributor_author_royalty_rate),
            "publication_date": book.publication_date.strftime("%Y-%m"),
            "date": sale_date_str,
            "sale_source": "distributor",
            "distributor": parser.DISTRIBUTOR_NAME,
            "author_paid": False,
        }


# ──────────────────────────────────────────────────────────────────────────────
# Concrete sheet handlers
# ──────────────────────────────────────────────────────────────────────────────

class _PrintRoyaltySheetHandler(_SheetHandler):
    """
    Handles "Paperback Royalty" and "Hardcover Royalty" sheets.

    Books are identified by ISBN (isbn_13 or isbn_10).
    Sale format: "print".
    Required columns: Title, Author, ISBN, Marketplace, Units Sold,
                      Units Refunded, Currency, Royalty (others ignored).
    """

    _REQUIRED_COLS = {
        "Title", "Author", "ISBN", "Marketplace",
        "Units Sold", "Units Refunded", "Currency", "Royalty",
    }

    def __init__(self, sheet_name: str) -> None:
        # Accepts either "Paperback Royalty" or "Hardcover Royalty"
        self.sheet_name = sheet_name

    def parse_rows(self, worksheet, filename, timestamp, parser):
        month_year, col_map, errors = self._extract_header(worksheet)
        if errors:
            return [], errors, []

        col_errors = self._check_required_cols(col_map, self._REQUIRED_COLS)
        if col_errors:
            return [], col_errors, []

        month, year = month_year
        sale_date_str = f"{year}-{month:02d}"
        preview_rows = []

        for row_num, row in enumerate(
            worksheet.iter_rows(min_row=3, values_only=True), start=3
        ):
            if _is_blank_row(row):
                continue

            row_ref = f"Sheet '{self.sheet_name}', row {row_num}"
            row_errors = []

            # Units Refunded must be zero
            refunded_raw = _cell(row, col_map, "Units Refunded")
            try:
                refunded = int(refunded_raw) if refunded_raw else None
                if refunded is None:
                    row_errors.append(f"{row_ref}: Units Refunded cannot be empty.")
                elif refunded != 0:
                    row_errors.append(
                        f"{row_ref}: Units Refunded must be zero (got {refunded}); "
                        f"rows with refunds cannot be imported."
                    )
            except (ValueError, TypeError):
                row_errors.append(
                    f"{row_ref}: Units Refunded '{refunded_raw}' is not a valid integer."
                )

            # Units Sold
            units_sold_raw = _cell(row, col_map, "Units Sold")
            units_sold = None
            try:
                units_sold = int(units_sold_raw) if units_sold_raw else None
                if units_sold is None:
                    row_errors.append(f"{row_ref}: Units Sold cannot be empty.")
            except (ValueError, TypeError):
                row_errors.append(
                    f"{row_ref}: Units Sold '{units_sold_raw}' is not a valid integer."
                )

            # ISBN → book lookup
            isbn_raw = _cell(row, col_map, "ISBN")
            book = None
            if not isbn_raw:
                row_errors.append(f"{row_ref}: ISBN is missing.")
            else:
                book = parser._lookup_book_by_isbn(isbn_raw)
                if not book:
                    row_errors.append(
                        f"{row_ref}: No book found with ISBN '{isbn_raw}'."
                    )

            # Royalty + currency conversion
            currency_raw = _cell(row, col_map, "Currency")
            royalty_raw = _cell(row, col_map, "Royalty")
            royalty_original, royalty_usd, conv_errors = self._convert_royalty(
                royalty_raw, currency_raw, row_ref, parser
            )
            row_errors.extend(conv_errors)

            if row_errors:
                errors.extend(row_errors)
                continue

            # Serializer validation
            currency = currency_raw.strip().upper()
            payload = {
                "book": book.id,
                "date": sale_date_str,
                "quantity": units_sold,
                "sale_source": "distributor",
                "distributor": parser.DISTRIBUTOR_NAME,
                "format": "print",
                "currency": currency,
                "publisher_revenue": str(royalty_usd),
            }
            if currency != "USD":
                payload["publisher_revenue_original"] = str(parser._money(royalty_original))

            validated_data, serial_errors = parser._validate_with_serializer(payload, row_ref)
            if serial_errors:
                errors.extend(serial_errors)
                continue

            # Build preview row
            marketplace = _cell(row, col_map, "Marketplace")
            author_royalty = parser._compute_author_royalty(book, royalty_usd)
            row_dict = self._preview_row_base(book, sale_date_str, parser)
            row_dict.update({
                "quantity": validated_data["quantity"],
                "kenp": None,
                "format": "print",
                "currency": currency,
                "publisher_revenue": str(parser._money(royalty_usd)),
                "publisher_revenue_original": payload.get("publisher_revenue_original"),
                "author_royalty": str(author_royalty),
                "comment": self._make_comment(marketplace, currency, filename, timestamp),
            })
            preview_rows.append(row_dict)

        if errors:
            return [], errors, []
        return preview_rows, [], []


class _EBookRoyaltySheetHandler(_SheetHandler):
    """
    Handles the "eBook Royalty" sheet.

    Books are identified by Amazon ebook ASIN.
    Sale format: "ebook".
    Required columns: Title, Author, ASIN, Marketplace, Units Sold,
                      Units Refunded, Net Units Sold, Currency, Royalty (others ignored).
    """

    sheet_name = "eBook Royalty"

    _REQUIRED_COLS = {
        "Title", "Author", "ASIN", "Marketplace",
        "Units Sold", "Units Refunded", "Net Units Sold", "Currency", "Royalty",
    }

    def parse_rows(self, worksheet, filename, timestamp, parser):
        month_year, col_map, errors = self._extract_header(worksheet)
        if errors:
            return [], errors, []

        col_errors = self._check_required_cols(col_map, self._REQUIRED_COLS)
        if col_errors:
            return [], col_errors, []

        month, year = month_year
        sale_date_str = f"{year}-{month:02d}"
        preview_rows = []

        for row_num, row in enumerate(
            worksheet.iter_rows(min_row=3, values_only=True), start=3
        ):
            if _is_blank_row(row):
                continue

            row_ref = f"Sheet 'eBook Royalty', row {row_num}"
            row_errors = []

            # Units Refunded must be zero
            refunded_raw = _cell(row, col_map, "Units Refunded")
            try:
                refunded = int(refunded_raw) if refunded_raw else None
                if refunded is None:
                    row_errors.append(f"{row_ref}: Units Refunded cannot be empty.")
                elif refunded != 0:
                    row_errors.append(
                        f"{row_ref}: Units Refunded must be zero (got {refunded}); "
                        f"rows with refunds cannot be imported."
                    )
            except (ValueError, TypeError):
                row_errors.append(
                    f"{row_ref}: Units Refunded '{refunded_raw}' is not a valid integer."
                )

            # Units Sold and Net Units Sold must match
            units_sold_raw = _cell(row, col_map, "Units Sold")
            net_units_raw = _cell(row, col_map, "Net Units Sold")
            units_sold = None
            net_units = None

            try:
                units_sold = int(units_sold_raw) if units_sold_raw else None
                if units_sold is None:
                    row_errors.append(f"{row_ref}: Units Sold cannot be empty.")
            except (ValueError, TypeError):
                row_errors.append(
                    f"{row_ref}: Units Sold '{units_sold_raw}' is not a valid integer."
                )

            try:
                net_units = int(net_units_raw) if net_units_raw else None
                if net_units is None:
                    row_errors.append(f"{row_ref}: Net Units Sold cannot be empty.")
            except (ValueError, TypeError):
                row_errors.append(
                    f"{row_ref}: Net Units Sold '{net_units_raw}' is not a valid integer."
                )

            if units_sold is not None and net_units is not None and units_sold != net_units:
                row_errors.append(
                    f"{row_ref}: Units Sold ({units_sold}) does not equal "
                    f"Net Units Sold ({net_units})."
                )

            # ASIN → book lookup
            asin_raw = _cell(row, col_map, "ASIN")
            book = None
            if not asin_raw:
                row_errors.append(f"{row_ref}: ASIN is missing.")
            else:
                book = parser._lookup_book_by_asin(asin_raw)
                if not book:
                    row_errors.append(
                        f"{row_ref}: No book found with ASIN '{asin_raw}'."
                    )

            # Royalty + currency conversion
            currency_raw = _cell(row, col_map, "Currency")
            royalty_raw = _cell(row, col_map, "Royalty")
            royalty_original, royalty_usd, conv_errors = self._convert_royalty(
                royalty_raw, currency_raw, row_ref, parser
            )
            row_errors.extend(conv_errors)

            if row_errors:
                errors.extend(row_errors)
                continue

            # Serializer validation
            currency = currency_raw.strip().upper()
            payload = {
                "book": book.id,
                "date": sale_date_str,
                "quantity": net_units,
                "sale_source": "distributor",
                "distributor": parser.DISTRIBUTOR_NAME,
                "format": "ebook",
                "currency": currency,
                "publisher_revenue": str(royalty_usd),
            }
            if currency != "USD":
                payload["publisher_revenue_original"] = str(parser._money(royalty_original))

            validated_data, serial_errors = parser._validate_with_serializer(payload, row_ref)
            if serial_errors:
                errors.extend(serial_errors)
                continue

            # Build preview row
            marketplace = _cell(row, col_map, "Marketplace")
            author_royalty = parser._compute_author_royalty(book, royalty_usd)
            row_dict = self._preview_row_base(book, sale_date_str, parser)
            row_dict.update({
                "quantity": validated_data["quantity"],
                "kenp": None,
                "format": "ebook",
                "currency": currency,
                "publisher_revenue": str(parser._money(royalty_usd)),
                "publisher_revenue_original": payload.get("publisher_revenue_original"),
                "author_royalty": str(author_royalty),
                "comment": self._make_comment(marketplace, currency, filename, timestamp),
            })
            preview_rows.append(row_dict)

        if errors:
            return [], errors, []
        return preview_rows, [], []


class _KENPSheetHandler(_SheetHandler):
    """
    Handles the "KENP" sheet (Kindle Unlimited and Audible revenue).

    Only rows where "eBook ASIN" is not "N/A" are supported; other rows
    (audiobook rows) generate a warning and are skipped.
    Books are identified by "eBook ASIN".
    Sale format: "kindle unlimited".
    Required columns: Title, Author, eBook ASIN, Marketplace,
                      Kindle Edition Normalized Pages (KENP), Currency, Royalty (others ignored).
    """

    sheet_name = "KENP"

    _REQUIRED_COLS = {
        "Title", "Author", "eBook ASIN", "Marketplace",
        "Kindle Edition Normalized Pages (KENP)", "Currency", "Royalty",
    }

    def parse_rows(self, worksheet, filename, timestamp, parser):
        month_year, col_map, errors = self._extract_header(worksheet)
        if errors:
            return [], errors, []

        col_errors = self._check_required_cols(col_map, self._REQUIRED_COLS)
        if col_errors:
            return [], col_errors, []

        month, year = month_year
        sale_date_str = f"{year}-{month:02d}"
        preview_rows = []
        warnings = []

        for row_num, row in enumerate(
            worksheet.iter_rows(min_row=3, values_only=True), start=3
        ):
            if _is_blank_row(row):
                continue

            row_ref = f"Sheet 'KENP', row {row_num}"
            asin_raw = _cell(row, col_map, "eBook ASIN")

            # Skip audiobook/Audible rows where ASIN is "N/A" or absent
            if not asin_raw or asin_raw.upper() == "N/A":
                warnings.append(
                    f"{row_ref}: Skipped — eBook ASIN is '{asin_raw or 'empty'}' "
                    f"(audiobook/Audible row not supported)."
                )
                continue

            row_errors = []

            # KENP page count
            kenp_raw = _cell(row, col_map, "Kindle Edition Normalized Pages (KENP)")
            kenp = None
            try:
                kenp = int(kenp_raw) if kenp_raw else None
                if kenp is None:
                    row_errors.append(f"{row_ref}: KENP cannot be empty.")
            except (ValueError, TypeError):
                row_errors.append(
                    f"{row_ref}: KENP '{kenp_raw}' is not a valid integer."
                )

            # ASIN → book lookup
            book = parser._lookup_book_by_asin(asin_raw)
            if not book:
                row_errors.append(
                    f"{row_ref}: No book found with eBook ASIN '{asin_raw}'."
                )

            # Royalty + currency conversion
            currency_raw = _cell(row, col_map, "Currency")
            royalty_raw = _cell(row, col_map, "Royalty")
            royalty_original, royalty_usd, conv_errors = self._convert_royalty(
                royalty_raw, currency_raw, row_ref, parser
            )
            row_errors.extend(conv_errors)

            if row_errors:
                errors.extend(row_errors)
                continue

            # Serializer validation (quantity absent — KU rows have kenp instead)
            currency = currency_raw.strip().upper()
            payload = {
                "book": book.id,
                "date": sale_date_str,
                "kenp": kenp,
                "sale_source": "distributor",
                "distributor": parser.DISTRIBUTOR_NAME,
                "format": "kindle unlimited",
                "currency": currency,
                "publisher_revenue": str(royalty_usd),
            }
            if currency != "USD":
                payload["publisher_revenue_original"] = str(parser._money(royalty_original))

            validated_data, serial_errors = parser._validate_with_serializer(payload, row_ref)
            if serial_errors:
                errors.extend(serial_errors)
                continue

            # Build preview row
            marketplace = _cell(row, col_map, "Marketplace")
            author_royalty = parser._compute_author_royalty(book, royalty_usd)
            row_dict = self._preview_row_base(book, sale_date_str, parser)
            row_dict.update({
                "quantity": None,
                "kenp": validated_data["kenp"],
                "format": "kindle unlimited",
                "currency": currency,
                "publisher_revenue": str(parser._money(royalty_usd)),
                "publisher_revenue_original": payload.get("publisher_revenue_original"),
                "author_royalty": str(author_royalty),
                "comment": self._make_comment(marketplace, currency, filename, timestamp),
            })
            preview_rows.append(row_dict)

        if errors:
            return [], errors, warnings
        return preview_rows, [], warnings


class _AudiobookRoyaltySheetHandler(_SheetHandler):
    """
    Handles the "Audiobook Royalty" sheet.

    This sheet is not supported for import. The handler only validates
    that no data rows are present; if rows exist, a warning is issued.
    """

    sheet_name = "Audiobook Royalty"

    def parse_rows(self, worksheet, filename, timestamp, parser):
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            if not _is_blank_row(row):
                return [], [], [
                    "Sheet 'Audiobook Royalty' contains data rows that will not be imported. "
                    "Audiobook royalties are not currently supported."
                ]
        return [], [], []


# ──────────────────────────────────────────────────────────────────────────────
# Top-level parser
# ──────────────────────────────────────────────────────────────────────────────

#: Sheet handlers for all sheets the parser should attempt to process.
#: Ordered: print sheets first, then ebook, then KENP, then audiobook check.
_SHEET_HANDLERS: list[_SheetHandler] = [
    _PrintRoyaltySheetHandler("Paperback Royalty"),
    _PrintRoyaltySheetHandler("Hardcover Royalty"),
    _EBookRoyaltySheetHandler(),
    _KENPSheetHandler(),
    _AudiobookRoyaltySheetHandler(),
]

#: Names of sheets that count as "supported data sheets" (at least one must be present).
_SUPPORTED_DATA_SHEET_NAMES = {
    "Paperback Royalty",
    "Hardcover Royalty",
    "eBook Royalty",
    "KENP",
}


class AmazonXLSXParser(SalesImportParser):
    """
    Parser for Amazon monthly sales XLSX files.

    The XLSX contains multiple named sheets. This parser:
      1. Opens the XLSX with openpyxl.
      2. Verifies at least one supported data sheet is present.
      3. Dispatches each present sheet to its _SheetHandler.
      4. Merges all preview rows, errors, and warnings into a single ParseResult.

    Month/year is extracted from each sheet's special first row (cell A1 = "Sales Period",
    cell B1 = e.g. "June 2025") — the user does NOT provide it.

    Currency conversion is handled by self._converter (a CurrencyConverter instance).
    Sheet handlers access it via parser._converter.to_usd(amount, currency).
    """

    DISTRIBUTOR_NAME = "Amazon"

    def __init__(self) -> None:
        self._converter = CurrencyConverter()

    def parse_and_validate(self, file, **kwargs) -> ParseResult:
        """
        Parse and validate an Amazon XLSX file.

        Args:
            file: An UploadedFile (from request.FILES).

        Returns:
            ParseResult. errors blocks preview; warnings are non-blocking.
        """
        filename = file.name or "unknown"
        timestamp = self._capture_timestamp()
        metadata = {"filename": filename, "timestamp": timestamp}

        # ---- open XLSX ----
        try:
            workbook = openpyxl.load_workbook(file, data_only=True)
        except Exception:
            return ParseResult(
                errors=["File is not a valid XLSX."],
                metadata=metadata,
            )

        present_sheet_names = set(workbook.sheetnames)

        # ---- at least one supported data sheet must be present ----
        if not (present_sheet_names & _SUPPORTED_DATA_SHEET_NAMES):
            return ParseResult(
                errors=[
                    "No supported sheets found. Expected at least one of: "
                    + ", ".join(sorted(_SUPPORTED_DATA_SHEET_NAMES)) + "."
                ],
                metadata=metadata,
            )

        # ---- dispatch to sheet handlers ----
        all_preview: list[dict] = []
        all_errors: list[str] = []
        all_warnings: list[str] = []

        for handler in _SHEET_HANDLERS:
            if handler.sheet_name not in present_sheet_names:
                continue
            ws = workbook[handler.sheet_name]
            rows, errors, warnings = handler.parse_rows(ws, filename, timestamp, self)
            all_preview.extend(rows)
            all_errors.extend(errors)
            all_warnings.extend(warnings)

        if all_errors:
            return ParseResult(errors=all_errors, warnings=all_warnings, metadata=metadata)

        return ParseResult(preview=all_preview, warnings=all_warnings, metadata=metadata)
